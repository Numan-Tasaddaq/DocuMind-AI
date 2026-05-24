from io import BytesIO

from openpyxl import load_workbook
from pypdf import PdfReader
from xlrd import open_workbook


class DocumentParserError(Exception):
    pass


def _parse_pdf(file_bytes: bytes) -> str:
    reader = PdfReader(BytesIO(file_bytes))
    pages: list[str] = []
    for page in reader.pages:
        pages.append(page.extract_text() or "")
    return "\n".join(pages).strip()


def _parse_docx(file_bytes: bytes) -> str:
    from docx import Document

    doc = Document(BytesIO(file_bytes))
    paragraphs = [paragraph.text for paragraph in doc.paragraphs if paragraph.text]
    return "\n".join(paragraphs).strip()


def _parse_xlsx(file_bytes: bytes) -> str:
    workbook = load_workbook(filename=BytesIO(file_bytes), data_only=True, read_only=True)
    blocks: list[str] = []
    for sheet in workbook.worksheets:
        blocks.append(f"[Sheet: {sheet.title}]")
        for row in sheet.iter_rows(values_only=True):
            row_values = [str(cell).strip() for cell in row if cell is not None and str(cell).strip()]
            if row_values:
                blocks.append(" | ".join(row_values))
    return "\n".join(blocks).strip()


def _parse_xls(file_bytes: bytes) -> str:
    workbook = open_workbook(file_contents=file_bytes)
    blocks: list[str] = []
    for sheet in workbook.sheets():
        blocks.append(f"[Sheet: {sheet.name}]")
        for row_index in range(sheet.nrows):
            row_values = []
            for col_index in range(sheet.ncols):
                value = sheet.cell_value(row_index, col_index)
                text_value = str(value).strip()
                if text_value:
                    row_values.append(text_value)
            if row_values:
                blocks.append(" | ".join(row_values))
    return "\n".join(blocks).strip()


def parse_document(file_bytes: bytes, extension: str) -> str:
    normalized = extension.lower().lstrip(".")

    if normalized == "pdf":
        text = _parse_pdf(file_bytes)
    elif normalized == "docx":
        text = _parse_docx(file_bytes)
    elif normalized == "xlsx":
        text = _parse_xlsx(file_bytes)
    elif normalized == "xls":
        text = _parse_xls(file_bytes)
    elif normalized == "doc":
        raise DocumentParserError("Legacy .doc is not supported yet. Please upload a .docx file.")
    else:
        raise DocumentParserError("Unsupported file format.")

    if not text:
        raise DocumentParserError("No extractable text found in document.")
    return text
